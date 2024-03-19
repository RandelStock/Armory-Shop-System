#=================================================IMPORTED CUSTOMTKINTER THEME======================================================================
#To Instal Customtkinter Type this in cmd / visual studio code terminal - pip install customtkinter
import customtkinter as ctk
from tkinter import messagebox
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from PIL import Image, ImageTk
from tkinter import ttk

root = ctk.CTk()
root.geometry('600x600')
root.title('LogIN')
root._set_appearance_mode('dark')
root.resizable(False,False)

excel_con = Workbook()
#=================================================IMAGE======================================================================
user_image = ctk.CTkImage(dark_image=Image.open(r"User.png"), size=(50, 50))
pass_image = ctk.CTkImage(dark_image=Image.open(r"pass.png"), size=(50, 50))
logo_image = ctk.CTkImage(dark_image=Image.open(r"witcher.png"), size=(80, 80))
loginbg_image = ctk.CTkImage(dark_image=Image.open(r"loginbgs.png"), size=(600, 600))

loginbgframe_label = ctk.CTkLabel(root, image=loginbg_image, text="")
loginbgframe_label.place(x=0,y=0)

#=================================================REGISTER FUNCTION======================================================================
def registerfunction():
    reg = ctk.CTkToplevel()
    reg.geometry('600x400')
    reg.resizable(False,False)
    reg.configure(bg='dimgray')
    reg.title('REGISTER')
    root.withdraw()

#=================================================ACCOUNTS(EXCEL FOR THE REGISTERED ACCOUNTS)======================================================================
    excel_con = load_workbook('accounts.xlsx')
    excel_activate = excel_con.active

    def register():
        Found = False
        user = username_ntr.get()
        password = password_ntr.get()
        email = em_ntr.get()
        
        if user == "" and password == "" and email == "":
            messagebox.showinfo("Notification", "All fields are required", parent=root)
        elif user == "" or password == "" or email == "":
            messagebox.showinfo("Notification", "All fields are required", parent=root)
        elif len(password) < 8:
            messagebox.showerror("Notification", "Password must be at least 8 characters long!", parent=root)
        else:
            for each_cell in range(2, excel_activate.max_row + 1):
                if user == excel_activate['A' + str(each_cell)].value:
                    Found = True
                    break
                else:
                    Found = False
            if Found:
                messagebox.showerror("ERROR", "Account Already Created!!!")
            else:
                lastrow = str(excel_activate.max_row + 1)
                excel_activate['A' + lastrow] = user
                excel_activate['B' + lastrow] = password
                excel_activate['C' + lastrow] = em_ntr.get()
                excel_con.save('accounts.xlsx')
                reg.destroy()
                messagebox.showinfo("SUCCESS", "Account Successfully Created!!!")
                root.deiconify()

    def go_back_to_main(root, reg):
        reg.destroy()
        root.deiconify()
        messagebox.showinfo("Login", "Back To LogIn")

    reminder_label = ctk.CTkLabel(reg,
                                  text='Put Your Email To Retrieve Or Change Your Password', 
                                  width=50,
                                  font=('arial',20,'underline'))
    user_label = ctk.CTkLabel(reg, 
                              text="USERNAME:", 
                              width=10,
                              font=('arial',20))
    pass_label = ctk.CTkLabel(reg, 
                              text="PASSWORD:", 
                              width=10,
                              font=('arial',20))
    username_ntr = ctk.CTkEntry(reg, 
                                width=200,
                                font=('arial',20),
                                placeholder_text='•USERNAME')
    password_ntr = ctk.CTkEntry(reg, 
                                width=200, 
                                show="•",
                                font=('arial',20),
                                placeholder_text='•PASSWORD')
    em_label = ctk.CTkLabel(reg, 
                            text="EMAIL:", 
                            width=10,
                            font=('arial',20))
    em_ntr = ctk.CTkEntry(reg, 
                          width=200,
                          font=('arial',20),
                          placeholder_text='•EMAIL')
    registers = ctk.CTkButton(reg,
                            text="Register",
                            command=lambda: register(),
                            width=16,
                            font=('arial',30),
                            fg_color='#323232',
                            hover_color='black')

    user_label.place(x=120, 
                    y=100)
    username_ntr.place(x=270, 
                       y=100)
    pass_label.place(x=120, 
                    y=150)
    password_ntr.place(x=270, 
                       y=150)
    em_label.place(x=120,
                   y=200)
    em_ntr.place(x=270,
                 y=200)
    registers.place(x=230,
                    y=250)
    reminder_label.place(x=50,
                        y=320)

    reg.protocol("WM_DELETE_WINDOW", lambda: go_back_to_main(root, reg))
    reg.mainloop()

#=================================================LOG-IN FUNCTION======================================================================
def loginfunction():
#=================================================ACCOUNTS(EXCEL FOR THE REGISTERED ACCOUNTS)======================================================================
    excel_con = load_workbook('accounts.xlsx')
    excel_activate = excel_con.active
    user = username.get()
    passw = password.get()
    found = False

    if user == "" or passw == "":
        messagebox.showinfo("Notification", "All fields are required")
    elif len(passw) < 8:
        messagebox.showerror("Notification", "Wrong Password / Username!\nTry Again!")
    else:
        for each_cell in range(2, excel_activate.max_row + 1):
            if (username.get() == excel_activate['A' + str(each_cell)].value and password.get() == excel_activate['B' + str(each_cell)].value):
                found = True
                break
#=================================================SHOP INTERFACE======================================================================
        if found:
            shop = ctk.CTkToplevel()
            shop.geometry("1520x700")
            shop.title('Armorer')
            shop.resizable(False, False)
            root.withdraw()

#=================================================ORDERS(EXCEL FOR THE ORDERED ITEMS)======================================================================
            excel_con = load_workbook('orders.xlsx')
            excel_activate = excel_con.active

#=================================================CLEARS ENTRY TEXT======================================================================
            def clear_login_entry():
                username.delete(0,'end')
                password.delete(0,'end')

#=================================================USE TO GO BACK TO LOGIN BY EXITING======================================================================
            def go_back_to_main(root, shop):
                shop.destroy()
                root.deiconify()
                messagebox.showinfo("LogOUT NOTIFICATION", "Back To LogIn")

#=================================================IMAGE FOR THE SHOP INTERFACE BACKGROUND======================================================================
            bg_image = ctk.CTkImage(dark_image=Image.open(r"armorybg.png"), size=(1520, 700))
            bg = ctk.CTkLabel(shop,image=bg_image)
            bg.place(x=0,y=0)

#=================================================ITEMS INFORMATIONS (NEW TOPLEVEL FOR EACH GUN INFORMATIONS)======================================================================
            def glock44_info():
                glock = ctk.CTkToplevel()
                glock.geometry('600x400')
                glock.title('GLOCK 44 INFORMATION')
                glock.resizable(False, False)
                shop.withdraw()

#=================================================SHOW SELECTED FOR THE GUN INFORMATION GO THROUGH THE SELECTED ITEMS : IN SHOP INTERFACE======================================================================
                def show_selected():
                    global selected
                    global selected1
                    selected = glock_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=savage_image) 
                    
#=================================================EXIT FUNCTION FOR GUN INFORMATION AND SHOP INTERFACE WILL BE REOPEN======================================================================
                def exit_function():
                    glock.destroy()
                    shop.deiconify()

#=================================================IMAGES FOR GUN IMAGE , ADDTOCART AND EXIT ======================================================================
                glock_image = ctk.CTkImage(dark_image=Image.open(r"glock44.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

#=================================================LABELS , CHECKBUTTON , ADDTOCART ,EXIT VARIABLES======================================================================
                glock_label = ctk.CTkLabel(glock,
                                        text='',
                                        image=glock_image)
                
                glock_label.place(x=25,y=20)
                
                
                glock_price_label =ctk.CTkLabel(glock,
                                                text='₱21,279.99\nTo CheckOut',
                                                font=('courier',30,'underline'))
                glock_price_label.place(x=350,y=100)
                
                glock_price_info =ctk.CTkLabel(glock,
                                                text='Head to the range with the GLOCK 44 - G44 .22 LR Rimfire Pistol. \nThis semiautomatic pistol features a 10-round capacity, \nas well as a polymer grip.',
                                                font=('courier',15))
                glock_price_info.place(x=25,y=300)

                glock_var = ctk.StringVar()
                glockchk = ctk.CTkCheckBox(glock, 
                                        text="",
                                        variable=glock_var, 
                                        onvalue="Glock 44 - G44 .22 LR Rimfire Pistol\n₱21,279.99\n", 
                                        offvalue="")
                glockchk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(glock,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:show_selected())
                addtocart.place(x=430,y=180)
                
                exit = ctk.CTkButton(glock,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                glock.mainloop()

            def taurus_info():
                taurus = ctk.CTkToplevel()
                taurus.geometry('600x400')
                taurus.title('Taurus INFORMATION')
                taurus.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = taurus_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=taurus_image)
                    

                def exit_function():
                    taurus.destroy()
                    shop.deiconify()

                taurus_image = ctk.CTkImage(dark_image=Image.open(r"taurus.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    


                taurus_label = ctk.CTkLabel(taurus,
                                        text='',
                                        image=taurus_image)
                taurus_label.place(x=25,y=20)
                
                taurus_price_label =ctk.CTkLabel(taurus,
                                                text='₱15,679.99\nTo CheckOut',
                                                font=('courier',30,'underline'))
                taurus_price_label.place(x=350,y=100)

                taurus_price_info =ctk.CTkLabel(taurus,
                                                text='For a powerful self-defense firearm, get the Taurus G3C 9mm Pistol.\nThis compact everyday carry pistol has a single action with \nrestrike and a 12-round capacity,\nas well as a fixed front sight and drift adjustable rear sight.\nThis pistol features a Striker Block, manual safety, trigger safety and loaded chamber indicator.',
                                                font=('courier',10),anchor='w')
                taurus_price_info.place(x=25,y=300)

                taurus_var = ctk.StringVar()
                tauruschk = ctk.CTkCheckBox(taurus, 
                                        text="",
                                        variable=taurus_var, 
                                        onvalue="Taurus G3C 9mm Pistol\n₱15,679.99\n", 
                                        offvalue="")
                tauruschk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(taurus,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(taurus,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                taurus.mainloop()

            def ruger_info():
                ruger = ctk.CTkToplevel()
                ruger.geometry('600x400')
                ruger.title('Ruger INFORMATION')
                ruger.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = ruger_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=ruger_image)
                    

                def exit_function():
                    ruger.destroy()
                    shop.deiconify()

                ruger_image = ctk.CTkImage(dark_image=Image.open(r"ruger.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                ruger_label = ctk.CTkLabel(ruger,
                                        text='',
                                        image=ruger_image)
                ruger_label.place(x=25,y=20)
                
                ruger_price_label =ctk.CTkLabel(ruger,
                                                text='₱21,279.99\nTo CheckOut',
                                                font=('courier',30,'underline'))
                ruger_price_label.place(x=350,y=100)
                
                ruger_price_info =ctk.CTkLabel(ruger,
                                                text='Made with a glass-filled nylon frame,\nthe Ruger EC9S 9mm Pistol includes a polymer grip\n and a double-action trigger.The pistol utilizes\n fixed sights and has a 7+1-round capacity.',
                                                font=('courier',15))
                ruger_price_info.place(x=60,y=300)

                ruger_var = ctk.StringVar()
                rugerchk = ctk.CTkCheckBox(ruger, 
                                        text="",
                                        variable=ruger_var, 
                                        onvalue="Ruger EC9S 9mm Pistol\n₱21,279.99\n", 
                                        offvalue="")
                rugerchk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(ruger,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(ruger,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                ruger.mainloop()

            def mete_info():
                mete = ctk.CTkToplevel()
                mete.geometry('600x400')
                mete.title('METE INFORMATION')
                mete.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = mete_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=mete_image)
                    

                def exit_function():
                    mete.destroy()
                    shop.deiconify()

                mete_image = ctk.CTkImage(dark_image=Image.open(r"mete.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                mete_label = ctk.CTkLabel(mete,
                                        text='',
                                        image=mete_image)
                mete_label.place(x=25,y=20)
                
                mete_price_label =ctk.CTkLabel(mete,
                                                text='₱21,279.99\nTo CheckOut',
                                                font=('courier',30,'underline'))
                mete_price_label.place(x=350,y=100)
                
                mete_price_info =ctk.CTkLabel(mete,
                                                text='The METE MC9 9mm 12RD Pistol is a semiautomatic,\nsubcompact pistol with a polymer frame and blacked-out\n dot sights. It accommodates 12+1 round and\n 15+1 round magazines, which are\n both included with the gun.',
                                                font=('courier',15))
                mete_price_info.place(x=50,y=300)

                mete_var = ctk.StringVar()
                metechk = ctk.CTkCheckBox(mete, 
                                        text="",
                                        variable=mete_var, 
                                        onvalue="METE MC9 9mm 12RD Pistol\n₱21,279.99\n", 
                                        offvalue="")
                metechk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(mete,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(mete,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                mete.mainloop()

            def charger_info():
                charger = ctk.CTkToplevel()
                charger.geometry('600x400')
                charger.title('RUGER CHARGER INFORMATION')
                charger.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = charger_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=rugercharger_image)

                def exit_function():
                    charger.destroy()
                    shop.deiconify()

                rugercharger_image = ctk.CTkImage(dark_image=Image.open(r"rugercharger.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                charger_label = ctk.CTkLabel(charger,
                                        text='',
                                        image=rugercharger_image)
                charger_label.place(x=25,y=20)
                
                charger_price_label =ctk.CTkLabel(charger,
                                                text='₱17,599.99\nTo CheckOut',
                                                font=('courier',30,'underline'))
                charger_price_label.place(x=350,y=100)
                
                charger_price_info =ctk.CTkLabel(charger,
                                                text='The Ruger Charger .22 LR Rimfire Pistol\n is a semiautomatic single-action pistol with a\n 15+1 capacity and a 10-inch threaded barrel.\nThe firearm also features a polymer stock.',
                                                font=('courier',15))
                charger_price_info.place(x=50,y=300)

                charger_var = ctk.StringVar()
                chargerchk = ctk.CTkCheckBox(charger, 
                                        text="",
                                        variable=charger_var, 
                                        onvalue="Ruger Charger .22\nLR Rimfire Pistol\n₱17,599.99\n", 
                                        offvalue="")
                chargerchk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(charger,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(charger,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                charger.mainloop()

            def sccy_info():
                sccy = ctk.CTkToplevel()
                sccy.geometry('600x400')
                sccy.title('SCCY INFORMATION')
                sccy.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = sccy_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=sccy_image)
                    

                def exit_function():
                    sccy.destroy()
                    shop.deiconify()

                sccy_image = ctk.CTkImage(dark_image=Image.open(r"sccy.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                sccy_label = ctk.CTkLabel(sccy,
                                        text='',
                                        image=sccy_image)
                sccy_label.place(x=25,y=20)
                
                sccy_price_label =ctk.CTkLabel(sccy,
                                                text='₱14,999.99\nTo CheckOut',
                                                font=('courier',30,'underline'))
                sccy_price_label.place(x=350,y=100)
                
                sccy_price_info =ctk.CTkLabel(sccy,
                                                text='Made with stainless steel and Zytel polymer,\nthe SCCY CPX-2 9mm Luger 2-Tone Pistol features\na 3.1-inch barrel and a grip with ergonomic finger grooves\nand a backstrap recoil cushion. The double-action trigger\nis safeguarded by a trigger guard key lock, and the 3-dot rear sight\nis adjustable for windage using the locking screw. Includes 2\ndouble-stack, 10-round magazines and 2 flat magazine bases.',
                                                font=('courier',10))
                sccy_price_info.place(x=100,y=300)

                sccy_var = ctk.StringVar()
                sccychk = ctk.CTkCheckBox(sccy, 
                                        text="",
                                        variable=sccy_var, 
                                        onvalue="SCCY CPX-2 9mm\nLuger 2-Tone Pistol\n₱14,999.99\n", 
                                        offvalue="")
                sccychk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(sccy,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(sccy,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                sccy.mainloop()

            def glock19_info():
                glock19 = ctk.CTkToplevel()
                glock19.geometry('600x400')
                glock19.title('GLOCK-19 INFORMATION')
                glock19.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = glock19_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=glock19_image)
                    

                def exit_function():
                    glock19.destroy()
                    shop.deiconify()

                glock19_image = ctk.CTkImage(dark_image=Image.open(r"glock19.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                glock19_label = ctk.CTkLabel(glock19,
                                        text='',
                                        image=glock19_image)
                glock19_label.place(x=25,y=20)
                
                glock19_price_label =ctk.CTkLabel(glock19,
                                                text='₱33,599.99\nTo CheckOut',
                                                font=('courier',30,'underline'))
                glock19_price_label.place(x=350,y=100)
                
                glock19_price_info =ctk.CTkLabel(glock19,
                                                text='The GLOCK 19 - G19X Crossover 9mm Luger Pistol\nis a double-action handgun with a 17+1 capacity.\nThe nitron-finished polymer frame is paired with a \npolygonal-rifled barrel and a Coyote nPVD slide.No manual safety.',
                                                font=('courier',15))
                glock19_price_info.place(x=25,y=300)

                glock19_var = ctk.StringVar()
                glock19chk = ctk.CTkCheckBox(glock19, 
                                        text="",
                                        variable=glock_var, 
                                        onvalue="GLOCK 19-G19X Gen5 NS 9mm\nCompact 17-Round Pistol\n₱33,599.99\n", 
                                        offvalue="")
                glock19chk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(glock19,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(glock19,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                glock19.mainloop()

            def taurusg2c_info():
                taurusg2c = ctk.CTkToplevel()
                taurusg2c.geometry('600x400')
                taurusg2c.title('TAURUS G2C INFORMATION')
                taurusg2c.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = taurusg2c_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=taurusg2c_image)
                    

                def exit_function():
                    taurusg2c.destroy()
                    shop.deiconify()

                taurusg2c_image = ctk.CTkImage(dark_image=Image.open(r"taurusg2c.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                taurusg2c_label = ctk.CTkLabel(taurusg2c,
                                        text='',
                                        image=taurusg2c_image)
                taurusg2c_label.place(x=25,y=20)
                
                taurusg2c_price_label =ctk.CTkLabel(taurusg2c,
                                                text='₱15,999.99\nTo CheckOut',
                                                font=('courier',30,'underline'))
                taurusg2c_price_label.place(x=350,y=100)
                
                taurusg2c_price_info =ctk.CTkLabel(taurusg2c,
                                                text='Designed with concealed carry in mind, the Taurus G2C SS\nRestrike 9mm Pistol boasts rugged, compact polymer frame with a streamlined,\nergonomic design and features a 3.2-inch barrel with a matte stainless finish.\nThis single-action firearm with Restrike capability is equipped\nwith a loaded chamber indicator, a manual safety and a fixed\nfront sight and adjustable rear sight.',
                                                font=('courier',12))
                taurusg2c_price_info.place(x=25,y=300)

                taurusg2c_var = ctk.StringVar()
                taurusg2cchk = ctk.CTkCheckBox(taurusg2c, 
                                        text="",
                                        variable=taurusg2c_var, 
                                        onvalue="Taurus G2C SS Restrike 9mm Pistol\n₱15,999.99\n", 
                                        offvalue="")
                taurusg2cchk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(taurusg2c,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(taurusg2c,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                taurusg2c.mainloop()

            def howa_info():
                howa = ctk.CTkToplevel()
                howa.geometry('600x400')
                howa.title('HOWA INFORMATION')
                howa.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = howa_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=howa_image)
                    

                def exit_function():
                    howa.destroy()
                    shop.deiconify()

                howa_image = ctk.CTkImage(dark_image=Image.open(r"howa.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                howa_label = ctk.CTkLabel(howa,
                                        text='',
                                        image=howa_image)
                howa_label.place(x=25,y=20)
                
                howa_price_label =ctk.CTkLabel(howa,
                                                text='₱78,400\nTo CheckOut',
                                                font=('courier',30,'underline'))
                howa_price_label.place(x=350,y=100)
                
                howa_price_info =ctk.CTkLabel(howa,
                                                text='The Howa Precision Chassis USA Flag 6.5 Creedmoor Rifle\nwith NS Diamond Scope is a bolt-action rifle with a\ncenterfire design and a Cerakote finish. It comes with\na 2-stage trigger, a 3-position safety and a long-range\nscope from Nikko Stirling.',
                                                font=('courier',15))
                howa_price_info.place(x=25,y=300)

                howa_var = ctk.StringVar()
                howachk = ctk.CTkCheckBox(howa, 
                                        text="",
                                        variable=howa_var, 
                                        onvalue="Howa Precision Chassis USA\nFlag 6.5 Creedmoor Rifle with\nNS Diamond Scope\n₱78,400\n", 
                                        offvalue="")
                howachk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(howa,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(howa,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                howa.mainloop()

            def savageaxis_info():
                savageaxis = ctk.CTkToplevel()
                savageaxis.geometry('600x400')
                savageaxis.title('Savage Axis INFORMATION')
                savageaxis.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = savageaxis_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=savageaxis_image)
                    

                def exit_function():
                    savageaxis.destroy()
                    shop.deiconify()

                savageaxis_image = ctk.CTkImage(dark_image=Image.open(r"savageaxis.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                savageaxis_label = ctk.CTkLabel(savageaxis,
                                        text='',
                                        image=savageaxis_image)
                savageaxis_label.place(x=25,y=20)
                
                savageaxis_price_label =ctk.CTkLabel(savageaxis,
                                                text='₱58,800\nTo CheckOut',
                                                font=('courier',30,'underline'))
                savageaxis_price_label.place(x=350,y=100)
                
                savageaxis_price_info =ctk.CTkLabel(savageaxis,
                                                text='Make the shot with the Savage Axis II Precision \nXP 6.5 Creedmoor Bolt-Action Rifle. The MDT chassis\nincludes an M-LOK fore-end and an MDT pistol grip,\nand the rifle features a user-adjustable AccuTrigger,\na 2-position tang safety and a\n1-piece 20 MOA rail. Made in USA.',
                                                font=('courier',15))
                savageaxis_price_info.place(x=25,y=300)

                savageaxis_var = ctk.StringVar()
                savageaxischk = ctk.CTkCheckBox(savageaxis, 
                                        text="",
                                        variable=savageaxis_var, 
                                        onvalue="Savage Axis II\nPrecision XP 6.5\nCreedmoorBolt-Action Rifle\n₱58,800\n", 
                                        offvalue="")
                savageaxischk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(savageaxis,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(savageaxis,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                savageaxis.mainloop()

            def barrett_info():
                barrett = ctk.CTkToplevel()
                barrett.geometry('600x400')
                barrett.title('BARRETT INFORMATION')
                barrett.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = barrett_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=barrett_image)
                    
                def exit_function():
                    barrett.destroy()
                    shop.deiconify()

                barrett_image = ctk.CTkImage(dark_image=Image.open(r"barrett.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                barrett_label = ctk.CTkLabel(barrett,
                                        text='',
                                        image=barrett_image)
                barrett_label.place(x=25,y=20)
                
                barrett_price_label =ctk.CTkLabel(barrett,
                                                text='₱966,000\nTo CheckOut',
                                                font=('courier',30,'underline'))
                barrett_price_label.place(x=350,y=100)
                
                barrett_price_info =ctk.CTkLabel(barrett,
                                                text='A limited number of Mk22 MRAD ASR Mod 0 rifle and optic kits\ncomplete with deployment case have been made available. \nThe .300 Norma is the new standard long range\nsniper round for the US Military.',
                                                font=('courier',15))
                barrett_price_info.place(x=25,y=300)

                barrett_var = ctk.StringVar()
                barrettchk = ctk.CTkCheckBox(barrett, 
                                        text="",
                                        variable=barrett_var, 
                                        onvalue="Barrett Mk22 Mod 0 ASR .300\nNorma-Nightforce SOCOM\nMilitary 7-35x Optics combo\n₱966,000\n", 
                                        offvalue="")
                barrettchk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(barrett,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(barrett,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                barrett.mainloop()

            def springfield_info():
                springfield = ctk.CTkToplevel()
                springfield.geometry('600x400')
                springfield.title('SPRINGFIELD INFORMATION')
                springfield.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = springfield_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=springfield_image)

                def exit_function():
                    springfield.destroy()
                    shop.deiconify()

                springfield_image = ctk.CTkImage(dark_image=Image.open(r"springfield.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                springfield_label = ctk.CTkLabel(springfield,
                                        text='',
                                        image=springfield_image)
                springfield_label.place(x=25,y=20)
                
                springfield_price_label =ctk.CTkLabel(springfield,
                                                text='₱92,344\nTo CheckOut',
                                                font=('courier',30,'underline'))
                springfield_price_label.place(x=350,y=100)
                
                springfield_price_info =ctk.CTkLabel(springfield,
                                                text='Chambered in 5.56mm, this semi-automatic rifle features a 16" Hammer\nForged 4150 Steel Barrel with a 1:7" twist rate. The 5-position\nadjustable stock features a cheek riser. Ambidextrous parts\ninclude the safety switch & charging handle. Each\nfirearm ships with one 30-round Magpul PMAG Magazine.',
                                                font=('courier',12))
                springfield_price_info.place(x=25,y=300)

                springfield_var = ctk.StringVar()
                springfieldchk = ctk.CTkCheckBox(springfield, 
                                        text="",
                                        variable=springfield_var, 
                                        onvalue="Springfield Hellion 5.56mm Semi\nAutomatic Fully Ambidextrous Bullpup\nFirstline Rifle with 16\n₱92,344\n", 
                                        offvalue="")
                springfieldchk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(springfield,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(springfield,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                springfield.mainloop()

            def ar15_info():
                ar15 = ctk.CTkToplevel()
                ar15.geometry('600x400')
                ar15.title('AR - 15 INFORMATION')
                ar15.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = ar15_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=ar15_image)
                    

                def exit_function():
                    ar15.destroy()
                    shop.deiconify()

                ar15_image = ctk.CTkImage(dark_image=Image.open(r"ar15.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                ar15_label = ctk.CTkLabel(ar15,
                                        text='',
                                        image=ar15_image)
                ar15_label.place(x=25,y=20)
                
                ar15_price_label =ctk.CTkLabel(ar15,
                                                text='₱30,800\nTo CheckOut',
                                                font=('courier',30,'underline'))
                ar15_price_label.place(x=350,y=100)
                
                ar15_price_info =ctk.CTkLabel(ar15,
                                                text=' Ready to hit the field or range out of the box, the LCAR comes equipped with a\ntrue,F-marked carbine-height front sight and folding Magpul backup rear sight.\nThe LCARs lightweight, 16” barrel is treated with a FNC hardening process\nfor improved longevity, corrosion resistance and precision accuracy exceeding\nmil-spec chrome plating. Mil-spec, six-position, M4 Stock, 7075 receivers and\nM16 bolt-carrier assembly complete a MSR package thats built to punch\npaper and bust game, all at a price set for any budget.',
                                                font=('courier',12))
                ar15_price_info.place(x=10,y=300)

                ar15_var = ctk.StringVar()
                ar15chk = ctk.CTkCheckBox(ar15, 
                                        text="",
                                        variable=ar15_var, 
                                        onvalue="DPMS LCAR 5.56mm Semi\nAutomatic AR-15 Rifle\n₱30,800\n", 
                                        offvalue="")
                ar15chk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(ar15,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(ar15,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                ar15.mainloop()

            def pof_info():
                pof = ctk.CTkToplevel()
                pof.geometry('600x400')
                pof.title('POF MINUTEMAN INFORMATION')
                pof.resizable(False, False)
                shop.withdraw()

                def show_selected():
                    global selected
                    global selected1
                    selected = pof_var.get()
                    selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                    selected1.pack()
                    selected1.configure(text=selected,image=pof_image)
                    

                def exit_function():
                    pof.destroy()
                    shop.deiconify()

                pof_image = ctk.CTkImage(dark_image=Image.open(r"pof.png"), size=(300, 300))
                addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(30, 30))
                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30))    

                pof_label = ctk.CTkLabel(pof,
                                        text='',
                                        image=pof_image)
                pof_label.place(x=25,y=20)
                
                pof_price_label =ctk.CTkLabel(pof,
                                                text='₱72,999.99\nTo CheckOut',
                                                font=('courier',30,'underline'))
                pof_price_label.place(x=350,y=100)
                
                pof_price_info =ctk.CTkLabel(pof,
                                                text='Minutemen were civilian colonists tasked with being a defensive force\nto be prepared at a moments notice. They took it upon\nthemselves to be trained in various weapons and tactics. We deeply\nadmire their patriotism and self-reliance. The “Minuteman”\nis our twist on what a direct impingement serious carbine should be,\ntopped with the controls and enhancements most\nimportant to remain mobile and light.',
                                                font=('courier',12))
                pof_price_info.place(x=25,y=300)

                pof_var = ctk.StringVar()
                pofchk = ctk.CTkCheckBox(pof, 
                                        text="",
                                        variable=pof_var, 
                                        onvalue="POF Minuteman 5.56 NATO Semi\nAutomatic Rifle with Tungsten\nCerakote Finish\n₱72,999.99\n", 
                                        offvalue="")
                pofchk.place(x=400,y=180)
                
                addtocart = ctk.CTkButton(pof,image=addtocart_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',command=lambda:show_selected())
                addtocart.place(x=430,y=180)

                exit = ctk.CTkButton(pof,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.place(x=550,y=0)
                    
                pof.mainloop()

#=================================================PISTOL,RIFLE AND SNIPER SEGMENTED BUTTON======================================================================
            def gun_button_callback(value):
                if gun_button_var.get() == 'Pistol':
#=================================================SCROLLABLE FRAME FOR ACCESSIBLE GUNS INFO======================================================================
                    scrollframe_gun = ctk.CTkScrollableFrame(shop,width=850,height=400,corner_radius=5)
                    scrollframe_gun.place(x=370,y=100)
#=================================================GUNS INFORMATION IMAGES ======================================================================
                    glock44 = ctk.CTkButton(scrollframe_gun,image=savage_image,text='',
                                fg_color='white',    
                                border_width=10,
                                command=lambda:glock44_info(),
                                border_color='black')
                    glock44.grid(row=0,
                                column=0,
                                padx=10,
                                pady=10)

                    glock44label= ctk.CTkLabel(scrollframe_gun,
                                        text='Glock 44 - G44 .22\nLR Rimfire Pistol',
                                        font=('courier',15,'underline'))
                    glock44label.grid(row=2,
                                    column=0,
                                    padx=10,
                                    pady=10)

                    taurus = ctk.CTkButton(scrollframe_gun,image=taurus_image,
                                        text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:taurus_info())
                    taurus.grid(row=0,
                                column=1,
                                padx=10,
                                pady=10)

                    tauruslabel= ctk.CTkLabel(scrollframe_gun,
                                        text='Taurus G3C 9mm Pistol',
                                        font=('courier',15,'underline'))
                    tauruslabel.grid(row=2,
                                    column=1,
                                    padx=10,
                                    pady=10)

                    ruger = ctk.CTkButton(scrollframe_gun,image=ruger_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:ruger_info())
                    ruger.grid(row=0,
                                column=2,
                                padx=10,
                                pady=10)

                    rugerlabel= ctk.CTkLabel(scrollframe_gun,
                                        text='Ruger EC9S 9mm Pistol',
                                        font=('courier',15,'underline'))
                    rugerlabel.grid(row=2,
                                    column=2,
                                    padx=10,
                                    pady=10)

                    mete = ctk.CTkButton(scrollframe_gun,image=mete_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:mete_info())
                    mete.grid(row=0,
                                column=3,
                                padx=10,
                                pady=10)

                    metelabel= ctk.CTkLabel(scrollframe_gun,
                                        text='METE MC9 9mm\n12RD Pistol',
                                        font=('courier',15,'underline'))
                    metelabel.grid(row=2,
                                    column=3,
                                    padx=10,
                                    pady=10)

                    charger = ctk.CTkButton(scrollframe_gun,image=rugercharger_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:charger_info())
                    charger.grid(row=3,
                                column=0,
                                padx=10,
                                pady=10)

                    chargerlabel= ctk.CTkLabel(scrollframe_gun,
                                        text='Ruger Charger .22\nLR Rimfire Pistol',
                                        font=('courier',15,'underline'))
                    chargerlabel.grid(row=4,
                                    column=0,
                                    padx=10,
                                    pady=10)

                    sccy = ctk.CTkButton(scrollframe_gun,image=sccy_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:sccy_info())
                    sccy.grid(row=3,
                                column=1,
                                padx=10,
                                pady=10)

                    sccylabel= ctk.CTkLabel(scrollframe_gun,
                                        text='SCCY CPX-2 9mm \nLuger 2-Tone Pistol',
                                        font=('courier',15,'underline'))
                    sccylabel.grid(row=4,
                                    column=1,
                                    padx=10,
                                    pady=10)

                    glock19 = ctk.CTkButton(scrollframe_gun,image=glock19_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:glock19_info())
                    glock19.grid(row=3,
                                column=2,
                                padx=10,
                                pady=10)

                    glock19label= ctk.CTkLabel(scrollframe_gun,
                                        text='GLOCK 19-G19X Gen5 NS 9mm\nCompact 17-Round Pistol',
                                        font=('courier',15,'underline'))
                    glock19label.grid(row=4,
                                    column=2,
                                    padx=10,
                                    pady=10)

                    taurusg2c = ctk.CTkButton(scrollframe_gun,image=taurusg2c_image,text='',
                                fg_color='white',
                                border_width=10,
                                border_color='black',
                                command=lambda:taurusg2c_info())
                    taurusg2c.grid(row=3,
                                column=3,
                                padx=10,
                                pady=10)

                    taurusg2clabel= ctk.CTkLabel(scrollframe_gun,
                                        text='Taurus G2C SS \nRestrike 9mm Pistol',
                                        font=('courier',15,'underline'))
                    taurusg2clabel.grid(row=4,
                                    column=3,
                                    padx=10,
                                    pady=10)
                elif gun_button_var.get() == 'Rifle':
                    scrollframe_gun = ctk.CTkScrollableFrame(shop,width=850,height=400,corner_radius=5)
                    scrollframe_gun.place(x=370,y=100)

                    springfield = ctk.CTkButton(scrollframe_gun,image=springfield_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:springfield_info())
                    springfield.grid(row=0,
                                column=0,
                                padx=10,
                                pady=10)

                    springfieldlabel= ctk.CTkLabel(scrollframe_gun,
                                        text='Springfield Hellion 5.56mm\nSemi-Automatic Fully-Ambidextrous\nBullpup Firstline Rifle with 16',
                                        font=('courier',15,'underline'))
                    springfieldlabel.grid(row=1,
                                    column=0,
                                    padx=10,
                                    pady=10)
                    
                    ar15 = ctk.CTkButton(scrollframe_gun,image=ar15_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:ar15_info())
                    ar15.grid(row=0,
                                column=1,
                                padx=10,
                                pady=10)

                    ar15label= ctk.CTkLabel(scrollframe_gun,
                                        text='DPMS LCAR 5.56mm\nSemi-Automatic AR-15 Rifle',
                                        font=('courier',15,'underline'))
                    ar15label.grid(row=1,
                                    column=1,
                                    padx=10,
                                    pady=10)
                    
                    pof = ctk.CTkButton(scrollframe_gun,image=pof_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:pof_info())
                    pof.grid(row=0,
                                column=2,
                                padx=10,
                                pady=10)

                    poflabel= ctk.CTkLabel(scrollframe_gun,
                                        text='POF Minuteman 5.56 NATO\nSemi-Automatic Rifle with\nTungsten Cerakote Finish',
                                        font=('courier',15,'underline'))
                    poflabel.grid(row=1,
                                    column=2,
                                    padx=10,
                                    pady=10)
                elif gun_button_var.get() == 'Sniper':
                    scrollframe_gun = ctk.CTkScrollableFrame(shop,width=850,height=400,corner_radius=5)
                    scrollframe_gun.place(x=370,y=100)

                    howa = ctk.CTkButton(scrollframe_gun,image=howa_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:howa_info())
                    howa.grid(row=0,
                                column=0,
                                padx=10,
                                pady=10)

                    howalabel= ctk.CTkLabel(scrollframe_gun,
                                        text='Howa Precision Chassis\nUSA Flag 6.5 Creedmoor Rifle\nwith NS Diamond Scope',
                                        font=('courier',15,'underline'))
                    howalabel.grid(row=1,
                                    column=0,
                                    padx=10,
                                    pady=10)
                    
                    savageaxis = ctk.CTkButton(scrollframe_gun,image=savageaxis_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:savageaxis_info())
                    savageaxis.grid(row=0,
                                column=1,
                                padx=10,
                                pady=10)

                    savageaxislabel= ctk.CTkLabel(scrollframe_gun,
                                        text='Savage Axis II\nPrecision XP 6.5 Creedmoor\nBolt-Action Rifle',
                                        font=('courier',15,'underline'))
                    savageaxislabel.grid(row=1,
                                    column=1,
                                    padx=10,
                                    pady=10)
                    
                    barrett = ctk.CTkButton(scrollframe_gun,image=barrett_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:barrett_info())
                    barrett.grid(row=0,
                                column=2,
                                padx=10,
                                pady=10)

                    barrettlabel= ctk.CTkLabel(scrollframe_gun,
                                        text='Barrett Mk22 Mod 0 ASR .300\nNorma-Nightforce SOCOM\nMilitary 7-35x Optics combo',
                                        font=('courier',15,'underline'))
                    barrettlabel.grid(row=1,
                                    column=2,
                                    padx=10,
                                    pady=10)

            gun_button_var = ctk.StringVar(value="")
            gun_button = ctk.CTkSegmentedButton(shop, values=["Pistol", "Rifle", "Sniper"],
                                                                command=gun_button_callback,
                                                                variable=gun_button_var,
                                                                font=('courier',32),
                                                                text_color='white',
                                                                selected_hover_color='black',
                                                                unselected_hover_color='black',
                                                                fg_color='black')

            gun_button.place(x=850,y=6)

#=================================================IMAGE PATHS======================================================================
            savage_image = ctk.CTkImage(dark_image=Image.open(r"glock44.png"), size=(150, 150))
            taurus_image = ctk.CTkImage(dark_image=Image.open(r"taurus.png"), size=(150, 150))
            ruger_image = ctk.CTkImage(dark_image=Image.open(r"C:ruger.png"), size=(150, 150))
            mete_image = ctk.CTkImage(dark_image=Image.open(r"mete.png"), size=(150, 150))
            rugercharger_image = ctk.CTkImage(dark_image=Image.open(r"rugercharger.png"), size=(150, 150))
            sccy_image = ctk.CTkImage(dark_image=Image.open(r"sccy.png"), size=(150, 150))
            glock19_image = ctk.CTkImage(dark_image=Image.open(r"glock19.png"), size=(150, 150))
            taurusg2c_image = ctk.CTkImage(dark_image=Image.open(r"taurusg2c.png"), size=(150, 150))
            howa_image = ctk.CTkImage(dark_image=Image.open(r"howa.png"), size=(200, 150))
            search_image = ctk.CTkImage(dark_image=Image.open(r"search.png"), size=(30, 30))
            savageaxis_image = ctk.CTkImage(dark_image=Image.open(r"savageaxis.png"), size=(200, 150))
            barrett_image = ctk.CTkImage(dark_image=Image.open(r"barrett.png"), size=(200, 150))
            springfield_image = ctk.CTkImage(dark_image=Image.open(r"springfield.png"), size=(200, 150))
            ar15_image = ctk.CTkImage(dark_image=Image.open(r"ar15.png"), size=(200, 150))
            pof_image = ctk.CTkImage(dark_image=Image.open(r"pof.png"), size=(200, 150))
            addtocart_image = ctk.CTkImage(dark_image=Image.open(r"addtocart.png"), size=(50, 50))
            checkout_image = ctk.CTkImage(dark_image=Image.open(r"checkout.png"), size=(50, 50))
            policy_image = ctk.CTkImage(dark_image=Image.open(r"policy.png"), size=(320, 160))
            banner_image = ctk.CTkImage(dark_image=Image.open(r"banner.png"), size=(850, 400))

#=================================================SCROLLABLE FRAME FOR ITEM LIST : (LEFT CORNER)======================================================================
            scrollframe = ctk.CTkScrollableFrame(shop,
                                                width=300,
                                                height=490,
                                                corner_radius=5)
            scrollframe.place(x=25,
                            y=25)

            scroll_label = ctk.CTkLabel(scrollframe,
                                        width=250,
                                        font=('Courier',30),
                                        text='ITEM LIST :',
                                        bg_color='#323232')
            scroll_label.grid(row=0,
                            column=0,
                            columnspan=4,
                            padx=10,
                            pady=10)

#=================================================FOR THE CHECKED BOXES PRINTED IN SELECTED ITEMS :======================================================================
            def show_selected():
                global selected
                checkboxes = [glock_var, taurus_var, ruger_var, mete_var,charger_var,sccy_var,glock19_var,taurusg2c_var,springfield_var,ar15_var,pof_var,howa_var,savageaxis_var,barrett_var]  
                selected = '\n'.join(var.get() for var in checkboxes)
                selected1.configure(text=selected)

#=================================================GUN CHECKBOX EVERY ITEM HAVE INDIVIDUAL STRINGVAR()======================================================================
            glock_var = ctk.StringVar()
            glockchk = ctk.CTkCheckBox(scrollframe, 
                                    text="Glock 44 - G44 .22 LR Rimfire Pistol",
                                    variable=glock_var, 
                                    onvalue=f"Glock 44 - G44 .22 LR Rimfire Pistol\n₱21,279.99\n", 
                                    offvalue="")
            glockchk.grid(row=1, column=0, padx=10, pady=10,stick='W')
            taurus_var = ctk.StringVar()
            tauruschk = ctk.CTkCheckBox(scrollframe, 
                                    text="Taurus G3C 9mm Pistol",
                                    variable=taurus_var, 
                                    onvalue="Taurus G3C 9mm Pistol\n₱15,679.99\n", 
                                    offvalue="")
            tauruschk.grid(row=2, column=0, padx=10, pady=10,stick='W')
            ruger_var = ctk.StringVar()
            rugerchk = ctk.CTkCheckBox(scrollframe, 
                                    text="Ruger EC9S 9mm Pistol",
                                    variable=ruger_var, 
                                    onvalue="Ruger EC9S 9mm Pistol\n ₱21,279.99\n", 
                                    offvalue="")
            rugerchk.grid(row=3, column=0, padx=10, pady=10,stick='W')
            mete_var = ctk.StringVar()
            metechk = ctk.CTkCheckBox(scrollframe, 
                                    text="METE MC9 9mm 12RD Pistol",
                                    variable=mete_var, 
                                    onvalue="METE MC9 9mm 12RD Pistol\n₱21,279.99", 
                                    offvalue="")
            metechk.grid(row=4, column=0, padx=10, pady=10,stick='W')
            charger_var = ctk.StringVar()
            chargerchk = ctk.CTkCheckBox(scrollframe, 
                                    text="The Ruger Charger .22 LR Rimfire Pistol",
                                    variable=charger_var, 
                                    onvalue="The Ruger Charger .22 LR Rimfire Pistol\n₱17,599.99\n", 
                                    offvalue="")
            chargerchk.grid(row=5, column=0, padx=10, pady=10,stick='W')
            sccy_var = ctk.StringVar()
            sccychk = ctk.CTkCheckBox(scrollframe, 
                                    text="SCCY CPX-2 9mm  Luger 2-Tone Pistol",
                                    variable=sccy_var, 
                                    onvalue="SCCY CPX-2 9mm  Luger 2-Tone Pistol\n₱14,999.99\n", 
                                    offvalue="")
            sccychk.grid(row=6, column=0, padx=10, pady=10,stick='W')
            glock19_var = ctk.StringVar()
            glock19chk = ctk.CTkCheckBox(scrollframe, 
                                    text="GLOCK 19-G19X Gen5 NS \n9mm Compact 17-Round Pistol",
                                    variable=glock19_var, 
                                    onvalue="GLOCK 19-G19X Gen5 NS \n9mm Compact 17-Round Pistol\n₱33,599.99\n", 
                                    offvalue="")
            glock19chk.grid(row=7, column=0, padx=10, pady=10,stick='W')
            taurusg2c_var = ctk.StringVar()
            taurusg2cchk = ctk.CTkCheckBox(scrollframe, 
                                    text="Taurus G2C SS Restrike 9mm Pistol",
                                    variable=taurusg2c_var, 
                                    onvalue="Taurus G2C SS Restrike 9mm Pistol\n₱15,999.99\n", 
                                    offvalue="")
            taurusg2cchk.grid(row=8, column=0, padx=10, pady=10,stick='W')
            springfield_var = ctk.StringVar()
            springfieldcchk = ctk.CTkCheckBox(scrollframe, 
                                    text="Springfield Hellion 5.56mm Semi-Automatic\nFully-Ambidextrous Bullpup\nFirstline Rifle with 16",
                                    variable=springfield_var, 
                                    onvalue="Springfield Hellion 5.56mm Semi-Automatic\nFully-Ambidextrous Bullpup\nFirstline Rifle with 16\n₱92,344\n", 
                                    offvalue="")
            springfieldcchk.grid(row=9, column=0, padx=10, pady=10,stick='W')
            ar15_var = ctk.StringVar()
            ar15cchk = ctk.CTkCheckBox(scrollframe, 
                                    text="DPMS LCAR 5.56mm Semi\nAutomatic AR-15 Rifle",
                                    variable=ar15_var, 
                                    onvalue="DPMS LCAR 5.56mm Semi\n-Automatic AR-15 Rifle\n₱30,800\n", 
                                    offvalue="")
            ar15cchk.grid(row=10, column=0, padx=10, pady=10,stick='W')
            pof_var = ctk.StringVar()
            pofcchk = ctk.CTkCheckBox(scrollframe, 
                                    text="POF Minuteman 5.56 NATO\nSemi-Automatic Rifle with\nTungsten Cerakote Finish",
                                    variable=pof_var, 
                                    onvalue="POF Minuteman 5.56 NATO\nSemi-Automatic Rifle with\nTungsten Cerakote Finish\n₱72,999.99\n", 
                                    offvalue="")
            pofcchk.grid(row=11, column=0, padx=10, pady=10,stick='W')
            howa_var = ctk.StringVar()
            howacchk = ctk.CTkCheckBox(scrollframe, 
                                    text="Howa Precision Chassis USA\nFlag 6.5 Creedmoor Rifle\nwith NS Diamond Scope",
                                    variable=howa_var, 
                                    onvalue="Howa Precision Chassis USA\nFlag 6.5 Creedmoor Rifle\nwith NS Diamond Scope\n₱78,400\n", 
                                    offvalue="")
            howacchk.grid(row=12, column=0, padx=10, pady=10,stick='W')
            savageaxis_var = ctk.StringVar()
            savageaxischk = ctk.CTkCheckBox(scrollframe, 
                                    text="Savage Axis II\nPrecision XP 6.5 Creedmoor\nBolt-Action Rifle",
                                    variable=savageaxis_var, 
                                    onvalue="Savage Axis II\nPrecision XP 6.5 Creedmoor\nBolt-Action Rifle\n₱58,800\n", 
                                    offvalue="")
            savageaxischk.grid(row=13, column=0, padx=10, pady=10,stick='W')
            barrett_var = ctk.StringVar()
            barrettchk = ctk.CTkCheckBox(scrollframe, 
                                    text="Barrett Mk22 Mod 0 ASR .300\nNorma-Nightforce SOCOM\nMilitary 7-35x Optics combo",
                                    variable=barrett_var, 
                                    onvalue="Barrett Mk22 Mod 0 ASR .300\nNorma-Nightforce SOCOM\nMilitary 7-35x Optics combo\n₱966,000\n", 
                                    offvalue="")
            barrettchk.grid(row=14, column=0, padx=10, pady=10,stick='W')

#=================================================SIMPLE POLICY IMAGE(BOTTOM LEFT CORNER)======================================================================
            policy_button = ctk.CTkLabel(shop, text='',image=policy_image, width=320,fg_color='#323232')
            policy_button.place(x=25, y=530)

#=================================================SELECT FRAME FOR THE SELECTED ITEMS TO BE PRINT ON(RIGHT CORNER)======================================================================
            select_frame = ctk.CTkScrollableFrame(shop,width=240,height=530)
            select_frame.place(x=1250,y=100)

            selected_label = ctk.CTkLabel(select_frame,text='SELECTED ITEMS :',font=('courier',25,'underline'))
            selected_label.pack()

            selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
            selected1.pack()

#=================================================BRAND NAME AT THE TOP======================================================================
            armorer_label = ctk.CTkLabel(shop,text='Welcome To ArmorerΩ',font=('courier',30,'underline'))
            armorer_label.place(x=370,y=10)

#=================================================GUN_FUNCTIONS IS LIST FOR THE SEARCH FUNCTION TO BASE ON======================================================================
            gun_functions = {
                'glock44': glock44_info,
                'taurus': taurus_info,
                'ruger': ruger_info,
                'mete':mete_info,
                'charger':charger_info,
                'sccy':sccy_info,
                'glock19':glock19_info,
                'taurusg2c':taurusg2c_info,
                'howa':howa_info,
                'savageaxis':savageaxis_info,
                'barrett':barrett_info,
                'springfield':springfield_info,
                'ar15':ar15_info,
                'pof':pof_info}

#=================================================SEARCH FUNCTION(EASY SEARCH AN ITEM EVEN IF ALL BIG LETTER)======================================================================
            def search_function():
                gun_list = ['glock44', 'taurus', 'ruger', 'mete', 'charger', 'sccy', 'glock19', 'taurusg2c', 'howa', 'savageaxis', 'barrett', 'springfield', 'ar15', 'pof']
                search_term = armorer_entry.get().lower()
                matching_gun = [name for name in gun_list if search_term in name.lower()]
                
                if matching_gun:
                    for gun in matching_gun:
                        if gun in gun_functions:
                            gun_functions[gun]()
                else:
                    messagebox.showinfo('Search Result', 'No matching items found.')

#=================================================SEARCH ORDER FUNCTION (FOR THE USE TO SEARCH AN ORDER IN THE EXCEL)======================================================================
            def search_order_function():
                so = ctk.CTkToplevel()
                so.geometry('500x80')
                so.title('SEARCH ORDER')
                so.resizable(False, False)
                shop.withdraw()

                exit_image = ctk.CTkImage(dark_image=Image.open(r"exit.png"), size=(30, 30)) 

                def exit_function():
                    so.destroy()
                    shop.deiconify()
                
                exit = ctk.CTkButton(so,image=exit_image,
                                    text='',
                                    width=10,
                                    fg_color='dimgray',
                                    command=lambda:exit_function())
                exit.pack(side='right',fill='y')

                def search_ord():
                    username = search_entry.get()

                    if username == "":
                        messagebox.showinfo('NOTIFICATION', 'USERNAME REQUIRED')
                    else:
                        Found = False
                        for each_cell in range(2, (excel_activate.max_row) + 1):
                            if username == excel_activate['A' + str(each_cell)].value:
                                Found = True
                                cell_address = str(each_cell)
                                break

                        if Found:
                            name = excel_activate['G' + str(each_cell)].value
                            ordered = excel_activate['B' + str(each_cell)].value
                            address = excel_activate['H' + str(each_cell)].value
                            message = f'DATA EXIST IN {cell_address}\nOrdered By -{name}\nOrdered Item - {ordered}\nWill Be Delivered At - {address}'
                            messagebox.showinfo("FOUND",message)
                        else:
                            messagebox.showinfo("NOT FOUND", "NO ORDER FOUND")

                search_entry = ctk.CTkEntry(so,width=200,font=('courier',30),placeholder_text='Enter Your Name')
                search_entry.pack(fill='x')

                search_btn = ctk.CTkButton(so,width=200,font=('courier',30),text='Enter',command=lambda:search_ord())
                search_btn.pack(fill='x')

                so.mainloop()

#=================================================ENTRY,BUTTONS FOR SEARCH/ADDTOCART/CHECKOUT/SEARCHORDER======================================================================
            armorer_entry = ctk.CTkEntry(shop, placeholder_text='Search Item', width=700, font=('courier', 30))
            armorer_entry.place(x=370, y=50)

            search_button = ctk.CTkButton(shop, image=search_image, text='Search Item', fg_color='dimgray', command=lambda:search_function(),hover_color='black',font=('courier',15))
            search_button.place(x=1070, y=50)

            addtocart_button = ctk.CTkButton(shop,image=addtocart_image,text='Add To\nCart',width=10,fg_color='dimgray',command=lambda:show_selected(),hover_color='black')
            addtocart_button.place(x=1250,y=25)

            checkout_button = ctk.CTkButton(shop,image=checkout_image,text='CheckOut',width=10,fg_color='dimgray',command=lambda:check_out(),hover_color='black')
            checkout_button.place(x=1380,y=25)

            search_order = ctk.CTkButton(shop,image=search_image,text='Search Order',width=50,fg_color='dimgray',command=lambda:search_order_function   (),hover_color='black',font=('courier',29))
            search_order.place(x=1250,y=650)

#=================================================FRAME FOR THE BANNER======================================================================
            scrollframe_gun = ctk.CTkFrame(shop,width=850,height=400,corner_radius=5)
            scrollframe_gun.place(x=370,y=100)

            banner = ctk.CTkLabel(scrollframe_gun,text='',image=banner_image)
            banner.pack()

#=================================================FRAME USE FOR THE TREEVIEW======================================================================
            tv_frame = ctk.CTkScrollableFrame(shop,width=850,height=3)
            tv_frame.place(x=370,y=520)

#=================================================STYLE FOR THE DESIGN OF THE TREEVIEW======================================================================
            style = ttk.Style()
            style.theme_use("clam")
            style.configure("Treeview", background="green", foreground="white", fieldbackground="#323232")

#=================================================TREEVIEW TO SEE THE DATA INSIDE OF THE EXCEL======================================================================
            tree = ttk.Treeview(tv_frame,height=100)
            tree['columns'] = ('Customer Name','Ordered Item','Location')

            tree.column('#0',width=0,stretch=0)
            tree.column('Customer Name',anchor='center',width=120)
            tree.column('Ordered Item',anchor='center',width=120)
            tree.column('Location',anchor='center',width=120)

            tree.heading('Customer Name',text='Customer Name')
            tree.heading('Ordered Item',text='Ordered Item')
            tree.heading('Location',text='Location')

            for each_cell in range(2, (excel_activate.max_row)+1):
                tree.insert(parent='', index="end", values=(excel_activate['A'+str(each_cell)].value,excel_activate['B'+str(each_cell)].value, excel_activate['H'+str(each_cell)].value))
            tree.pack(fill='x')

#=================================================CHECKOUT INTERFACE======================================================================
            def check_out():
                co = ctk.CTkToplevel()
                co.geometry('1000x700')
                co.title('CHECK OUT')
                co.resizable(False,False)
                shop.withdraw()

#=================================================BRAND NAME======================================================================
                co_logo = ctk.CTkLabel(co,text='ArmorerΩ',font=('courier',30,'underline','bold'))
                co_logo.place(x=25,y=5)

#=================================================TO DESTROY CHECKOUT INTERFACE TO GO BACK TO SHOP INTERFACE======================================================================
                def go_back_to_main(shop, co):
                    co.destroy()
                    shop.deiconify()
                    messagebox.showinfo("Shop", "Back To Shop")

#=================================================SEGMENTED BUTTON MAIN,PISTOL,RIFLE,SNIPER======================================================================
                def gun_button_callback(value):
                    if gun_button_var.get() == 'Main':
                        global co_frame
                        co_frame = ctk.CTkScrollableFrame(co, width=428, height=650, corner_radius=7)
                        co_frame.place(x=0, y=50)

                        co_label = ctk.CTkLabel(co_frame, text='CheckOut', font=('Arial', 50, 'underline', 'bold'))
                        co_label.grid(row=0,column=1)
                    elif gun_button_var.get() == 'Pistol':
                        def show_selected():
                            global selected
                            checkboxes = [glock44_var, taurus_var, ruger_var, mete_var, charger_var, sccy_var, glock19_var, taurusg2c_var] 
                            selected = '\n\n'.join(var.get() for var in checkboxes)
                            selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                            selected1.pack()
                            selected1.configure(text=selected)
                            for checkbox in checkboxes:
                                checkbox.set('')
                            
                        scrollframe_gun = ctk.CTkScrollableFrame(co,width=428,height=650,corner_radius=7)
                        scrollframe_gun.place(x=0,y=50)

                        glock44_var = ctk.StringVar()
                        glock44chk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=glock44_var, 
                                    onvalue="Glock 44-G44 .22\nLR Rimfire Pistol\n₱21,279.99\n", 
                                    offvalue="",
                                    text='₱21,279.99')
                        glock44chk.grid(row=0,
                                    column=0,
                                    padx=10,
                                    pady=10)
                        glock44 = ctk.CTkButton(scrollframe_gun,image=savage_image,text='',
                                    fg_color='white',    
                                    border_width=10,
                                    command=lambda:glock44_info(),
                                    border_color='black')
                        glock44.grid(row=0,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        glock44label= ctk.CTkLabel(scrollframe_gun,
                                            text='Glock 44 - G44 .22\nLR Rimfire Pistol',
                                            font=('courier',15,'underline'))
                        glock44label.grid(row=1,
                                        column=1,
                                        padx=10,
                                        pady=10)
                        
                        taurus_var = ctk.StringVar()
                        tauruschk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=taurus_var, 
                                    onvalue='Taurus G3C 9mm Pistol\n₱15,679.99\n', 
                                    offvalue="",
                                    text='₱15,679.99')
                        tauruschk.grid(row=2,
                                    column=0,
                                    padx=10,
                                    pady=10)
                        taurus = ctk.CTkButton(scrollframe_gun,image=taurus_image,
                                            text='',
                                            fg_color='white',
                                            border_width=10,
                                            border_color='black',
                                            command=lambda:taurus_info())
                        taurus.grid(row=2,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        tauruslabel= ctk.CTkLabel(scrollframe_gun,
                                            text='Taurus G3C 9mm Pistol',
                                            font=('courier',15,'underline'))
                        tauruslabel.grid(row=3,column=1,padx=10,pady=10)

                        ruger_var = ctk.StringVar()
                        rugerchk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=ruger_var, 
                                    onvalue='Ruger EC9S 9mm Pistol\n₱21,279.99\n', 
                                    offvalue="",
                                    text='₱21,279.99')
                        rugerchk.grid(row=4,
                                    column=0,
                                    padx=10,
                                    pady=10)
                        ruger = ctk.CTkButton(scrollframe_gun,image=ruger_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:ruger_info())
                        ruger.grid(row=4,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        rugerlabel= ctk.CTkLabel(scrollframe_gun,
                                            text='Ruger EC9S 9mm Pistol',
                                            font=('courier',15,'underline'))
                        rugerlabel.grid(row=5,
                                        column=1,
                                        padx=10,
                                        pady=10)

                        mete_var = ctk.StringVar()
                        metechk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=mete_var, 
                                    onvalue='METE MC9 9mm\n12RD Pistol\n₱21,279.99\n', 
                                    offvalue="",
                                    text='₱21,279.99')
                        metechk.grid(row=6,
                                    column=0,
                                    padx=10,
                                    pady=10)
                        mete = ctk.CTkButton(scrollframe_gun,image=mete_image,text='',
                                            fg_color='white',
                                            border_width=10,
                                            border_color='black',
                                            command=lambda:mete_info())
                        mete.grid(row=6,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        metelabel= ctk.CTkLabel(scrollframe_gun,
                                            text='METE MC9 9mm\n12RD Pistol',
                                            font=('courier',15,'underline'))
                        metelabel.grid(row=7,
                                        column=1,
                                        padx=10,
                                        pady=10)

                        charger_var = ctk.StringVar()
                        chargerchk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=charger_var, 
                                    onvalue='Ruger Charger .22\nLR Rimfire Pistol\n₱17,599.99\n', 
                                    offvalue="",
                                    text='₱17,599.99')
                        chargerchk.grid(row=8,
                                    column=0,
                                    padx=10,
                                    pady=10)
                        charger = ctk.CTkButton(scrollframe_gun,image=rugercharger_image,text='',
                                            fg_color='white',
                                            border_width=10,
                                            border_color='black',
                                            command=lambda:charger_info())
                        charger.grid(row=8,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        chargerlabel= ctk.CTkLabel(scrollframe_gun,
                                            text='Ruger Charger .22\nLR Rimfire Pistol',
                                            font=('courier',15,'underline'))
                        chargerlabel.grid(row=9,
                                        column=1,
                                        padx=10,
                                        pady=10)

                        sccy_var = ctk.StringVar()
                        sccychk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=sccy_var, 
                                    onvalue='SCCY CPX-2 9mm \nLuger 2-Tone Pistol\n₱14,999.99\n', 
                                    offvalue="",
                                    text='₱14,999.99')
                        sccychk.grid(row=10,
                                    column=0,
                                    padx=10,
                                    pady=10)
                        sccy = ctk.CTkButton(scrollframe_gun,image=sccy_image,text='',
                                            fg_color='white',
                                            border_width=10,
                                            border_color='black',
                                            command=lambda:sccy_info())
                        sccy.grid(row=10,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        sccylabel= ctk.CTkLabel(scrollframe_gun,
                                            text='SCCY CPX-2 9mm \nLuger 2-Tone Pistol',
                                            font=('courier',15,'underline'))
                        sccylabel.grid(row=11,
                                        column=1,
                                        padx=10,
                                        pady=10)

                        glock19_var = ctk.StringVar()
                        glock19chk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=glock19_var, 
                                    onvalue='GLOCK 19-G19X Gen5 NS 9mm\nCompact 17-Round Pistol\n₱33,599.99\n', 
                                    offvalue="",
                                    text='₱33,599.99')
                        glock19chk.grid(row=12,
                                    column=0,
                                    padx=10,
                                    pady=10)
                        glock19 = ctk.CTkButton(scrollframe_gun,image=glock19_image,text='',
                                            fg_color='white',
                                            border_width=10,
                                            border_color='black',
                                            command=lambda:glock19_info())
                        glock19.grid(row=12,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        glock19label= ctk.CTkLabel(scrollframe_gun,
                                            text='GLOCK 19-G19X Gen5 NS 9mm\nCompact 17-Round Pistol',
                                            font=('courier',15,'underline'))
                        glock19label.grid(row=13,
                                        column=1,
                                        padx=10,
                                        pady=10)

                        taurusg2c_var = ctk.StringVar()
                        taurusg2cchk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=taurusg2c_var, 
                                    onvalue='Taurus G2C SS \nRestrike 9mm Pistol\n₱15,999.99\n', 
                                    offvalue="",
                                    text='₱15,999.99')
                        taurusg2cchk.grid(row=14,
                                    column=0,
                                    padx=10,
                                    pady=10)
                        taurusg2c = ctk.CTkButton(scrollframe_gun,image=taurusg2c_image,text='',
                                    fg_color='white',
                                    border_width=10,
                                    border_color='black',
                                    command=lambda:taurusg2c_info())
                        taurusg2c.grid(row=14,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        taurusg2clabel= ctk.CTkLabel(scrollframe_gun,
                                            text='Taurus G2C SS \nRestrike 9mm Pistol',
                                            font=('courier',15,'underline'))
                        taurusg2clabel.grid(row=15,
                                        column=1,
                                        padx=10,
                                        pady=10)
                        
                        check_out_btn = ctk.CTkButton(scrollframe_gun,text='',image=addtocart_image,command=lambda:show_selected())
                        check_out_btn.grid(row=16,column=1,padx=10,pady=10)
                    elif gun_button_var.get() == 'Rifle':
                        scrollframe_gun = ctk.CTkScrollableFrame(co,width=428,height=650,corner_radius=7)
                        scrollframe_gun.place(x=0,y=50)
                        def show_selected():
                            global selected
                            checkboxes = [springfield_var,ar15_var,pof_var]  
                            selected = '\n\n'.join(var.get() for var in checkboxes)
                            selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                            selected1.pack()
                            selected1.configure(text=selected)
                            for checkbox in checkboxes:
                                checkbox.set('')

                        springfield_var = ctk.StringVar()
                        springfieldchk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=springfield_var, 
                                    onvalue='Springfield Hellion 5.56mm\nSemi-Automatic Fully-Ambidextrous\nBullpup Firstline Rifle with 16\n₱92,344\n', 
                                    offvalue="",
                                    text='₱92,344')
                        springfieldchk.grid(row=0,column=0)
                        springfield = ctk.CTkButton(scrollframe_gun,image=springfield_image,text='',
                                            fg_color='white',
                                            border_width=10,
                                            border_color='black',
                                            command=lambda:springfield_info())
                        springfield.grid(row=0,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        springfieldlabel= ctk.CTkLabel(scrollframe_gun,
                                            text='Springfield Hellion 5.56mm\nSemi-Automatic Fully-Ambidextrous\nBullpup Firstline Rifle with 16',
                                            font=('courier',15,'underline'))
                        springfieldlabel.grid(row=1,
                                        column=1,
                                        padx=10,
                                        pady=10)
                        
                        ar15_var = ctk.StringVar()
                        ar15chk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=ar15_var, 
                                    onvalue='DPMS LCAR 5.56mm\nSemi-Automatic AR-15 Rifle\n₱30,800\n', 
                                    offvalue="",
                                    text='₱30,800')
                        ar15chk.grid(row=2,column=0)
                        ar15 = ctk.CTkButton(scrollframe_gun,image=ar15_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:ar15_info())
                        ar15.grid(row=2,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        ar15label= ctk.CTkLabel(scrollframe_gun,
                                            text='DPMS LCAR 5.56mm\nSemi-Automatic AR-15 Rifle',
                                            font=('courier',15,'underline'))
                        ar15label.grid(row=3,
                                        column=1,
                                        padx=10,
                                        pady=10)
                        
                        pof_var = ctk.StringVar()
                        pofchk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=pof_var, 
                                    onvalue='POF Minuteman 5.56 NATO\nSemi-Automatic Rifle with\nTungsten Cerakote Finish\n₱72,999.99\n', 
                                    offvalue="",
                                    text='₱72,999.99')
                        pofchk.grid(row=4,column=0)
                        pof = ctk.CTkButton(scrollframe_gun,image=pof_image,text='',
                                            fg_color='white',
                                            border_width=10,
                                            border_color='black',
                                            command=lambda:pof_info())
                        pof.grid(row=4,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        poflabel= ctk.CTkLabel(scrollframe_gun,
                                            text='POF Minuteman 5.56 NATO\nSemi-Automatic Rifle with\nTungsten Cerakote Finish',
                                            font=('courier',15,'underline'))
                        poflabel.grid(row=5,
                                        column=1,
                                        padx=10,
                                        pady=10)
                        check_out_btn = ctk.CTkButton(scrollframe_gun,text='',image=addtocart_image,command=lambda:show_selected())
                        check_out_btn.grid(row=16,column=1,padx=10,pady=10)
                    elif gun_button_var.get() == 'Sniper':
                        scrollframe_gun = ctk.CTkScrollableFrame(co,width=428,height=650,corner_radius=7)
                        scrollframe_gun.place(x=0,y=50)
                        def show_selected():
                            global selected
                            checkboxes = [howa_var,savageaxis_var,barrett_var]  
                            selected = '\n\n'.join(var.get() for var in checkboxes)
                            selected1 = ctk.CTkLabel(select_frame, text='', font=('courier', 10))
                            selected1.pack()
                            selected1.configure(text=selected)
                            for checkbox in checkboxes:
                                checkbox.set('')

                        howa_var = ctk.StringVar()
                        howachk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=howa_var, 
                                    onvalue='Howa Precision Chassis\nUSA Flag 6.5 Creedmoor Rifle\nwith NS Diamond Scope\n₱78,400\n', 
                                    offvalue="",
                                    text='₱78,400')
                        howachk.grid(row=0,column=0)
                        howa = ctk.CTkButton(scrollframe_gun,image=howa_image,text='',
                                            fg_color='white',
                                            border_width=10,
                                            border_color='black',
                                            command=lambda:howa_info())
                        howa.grid(row=0,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        howalabel= ctk.CTkLabel(scrollframe_gun,
                                            text='Howa Precision Chassis\nUSA Flag 6.5 Creedmoor Rifle\nwith NS Diamond Scope',
                                            font=('courier',15,'underline'))
                        howalabel.grid(row=1,
                                        column=1,
                                        padx=10,
                                        pady=10)
                        
                        savageaxis_var = ctk.StringVar()
                        savageaxischk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=savageaxis_var, 
                                    onvalue='Savage Axis II\nPrecision XP 6.5 Creedmoor\nBolt-Action Rifle\n₱58,800\n', 
                                    offvalue="",
                                    text='₱58,800')
                        savageaxischk.grid(row=2,column=0)
                        savageaxis = ctk.CTkButton(scrollframe_gun,image=savageaxis_image,text='',
                                        fg_color='white',
                                        border_width=10,
                                        border_color='black',
                                        command=lambda:savageaxis_info())
                        savageaxis.grid(row=2,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        savageaxislabel= ctk.CTkLabel(scrollframe_gun,
                                            text='Savage Axis II\nPrecision XP 6.5 Creedmoor\nBolt-Action Rifle',
                                            font=('courier',15,'underline'))
                        savageaxislabel.grid(row=3,
                                        column=1,
                                        padx=10,
                                        pady=10)
                        
                        barrett_var = ctk.StringVar()
                        barrettchk = ctk.CTkCheckBox(scrollframe_gun,
                                    variable=barrett_var, 
                                    onvalue='Barrett Mk22 Mod 0 ASR .300\nNorma-Nightforce SOCOM\nMilitary 7-35x Optics combo\n₱966,000\n', 
                                    offvalue="",
                                    text='₱966,000')
                        barrettchk.grid(row=4,column=0)
                        barrett = ctk.CTkButton(scrollframe_gun,image=barrett_image,text='',
                                            fg_color='white',
                                            border_width=10,
                                            border_color='black',
                                            command=lambda:barrett_info())
                        barrett.grid(row=4,
                                    column=1,
                                    padx=10,
                                    pady=10)

                        barrettlabel= ctk.CTkLabel(scrollframe_gun,
                                            text='Barrett Mk22 Mod 0 ASR .300\nNorma-Nightforce SOCOM\nMilitary 7-35x Optics combo',
                                            font=('courier',15,'underline'))
                        barrettlabel.grid(row=5,
                                        column=1,
                                        padx=10,
                                        pady=10)
                        check_out_btn = ctk.CTkButton(scrollframe_gun,text='',image=addtocart_image,command=lambda:show_selected())
                        check_out_btn.grid(row=16,column=1,padx=10,pady=10)
                gun_button_var = ctk.StringVar(value="")
                gun_button = ctk.CTkSegmentedButton(co, values=["Main","Pistol", "Rifle", "Sniper"],
                                                                    command=gun_button_callback,
                                                                    variable=gun_button_var,
                                                                    fg_color='#323232',
                                                                    selected_hover_color='black',
                                                                    unselected_hover_color='black',
                                                                    font=('courier',15))

                gun_button.place(x=200,y=10)

#=================================================TO REFRESH AND UPDATE THE DATA IN THE EXCEL======================================================================
                def refresh_data(tree):
                    tree.delete(*tree.get_children())
                    data = get_updated_data()
                    for item in data:
                        tree.insert('', 'end', values=item)
                def get_updated_data():
                    
                    updated_value = list()
                    for each_cell in range(2, (excel_activate.max_row)+1):     
                        updated_value.append([excel_activate['A'+str(each_cell)].value,excel_activate['B'+str(each_cell)].value, excel_activate['H'+str(each_cell)].value])
                    return updated_value

#=================================================TO CLEAR THE ENTRIES======================================================================
                def clear():
                    credit_var.set('')
                    gcash_entry.delete(0,'end')
                    paymaya_entry.delete(0,'end')
                    username_entry.delete(0,'end')
                    name_entry.delete(0,'end')
                    add_entry.delete(0,'end')
                    glock_var.set('')
                    taurus_var.set('')
                    ruger_var.set('')
                    mete_var.set('')
                    charger_var.set('')
                    sccy_var.set('')
                    glock19_var.set('')
                    taurusg2c_var.set('')
                    springfield_var.set('')
                    ar15_var.set('')
                    pof_var.set('')
                    howa_var.set('')
                    savageaxis_var.set('')
                    barrett_var.set('')

#=================================================FOR THE PAY FUNCTIONS TO SAVE THE DATAS TO THE EXCEL======================================================================
                def show_selected(co_frame):
                            creditcard = credit_var.get()
                            gcash = gcash_entry.get()
                            paymaya = paymaya_entry.get()
                            username = username_entry.get()
                            full_name = name_entry.get()
                            address = add_entry.get()
                            Found = False
                            for each_cell in range(2, (excel_activate.max_row)+1):
                                if (username ==  excel_activate['A'+str(each_cell)].value):
                                    Found = True;
                                    break;
                                else:
                                    Found=False
                            if(Found == True):
                                messagebox.showerror("ERROR","USER ALREADY ORDER(POLICY)")
                            elif username == "":
                                messagebox.showerror("ERROR","USER REQUIRED")
                            else: 
                                checkout = ctk.CTkLabel(co_frame, text='', font=('courier', 20,'underline'))
                                checkout.grid(row=1,column=1)
                                checkout.configure(text=f'{selected}You Use Your : \n{creditcard}\n\nYour Gcash \n#{gcash}\n\nYour Paymaya \n#{paymaya}\n\nUserName :\n{username}\n\nFull Name :\n{full_name}\n\nDrop Location :\n{address}')
                                lastRow = str(excel_activate.max_row+1)
                                excel_activate['A'+lastRow] =  username
                                excel_activate['B'+lastRow] =  selected
                                excel_activate['D'+lastRow] =  creditcard
                                excel_activate['E'+lastRow] =  gcash
                                excel_activate['F'+lastRow] =  paymaya
                                excel_activate['G'+lastRow] =  full_name
                                excel_activate['H'+lastRow] =  address
                                messagebox.showinfo("SUCCESS","PAY SUCCESSFULL WAIT FOR YOUR ITEM TO ARRIVE\n THANK YOU!")
                                excel_con.save('orders.xlsx')
                                clear_selecteditems()
                                clear()
                                refresh_data(tree)

#=================================================TO CLEAR THE ORDER ITEM IN SELECTED ITEMS :======================================================================
                def clear_selecteditems():
                    selected1.configure(text="")

#=================================================TO DELETE THE DATA IN THE EXCEL======================================================================
                def cancel_order():
                    cancel = ctk.CTkToplevel()
                    cancel.geometry('500x100')
                    cancel.title('CANCEL ORDER')
                    cancel.resizable(False,False)
                    co.withdraw()
                    
                    def cancel_function():
                        for each_cell in range(2, (excel_activate.max_row)+1):
                            if (cancel_entry.get() ==  excel_activate['A'+str(each_cell)].value):
                                Found = True
                                cell_address = each_cell
                                break;
                            else:
                                Found=False
                        if(Found == True):
                            excel_activate.delete_rows(cell_address)
                            messagebox.showinfo("ORDER","ORDER CANCELLED")
                            excel_con.save('orders.xlsx')
                            refresh_data(tree)
                            clear_selecteditems()
                            clear()
                            cancel.destroy()
                            shop.deiconify()
                        else:
                            messagebox.showerror("ORDER","NO ORDER FOUND")
                    
                    cancel_entry = ctk.CTkEntry(cancel,placeholder_text='•ENTER YOUR USERNAME',font=('courier',30,'bold'),width=500)
                    cancel_entry.pack(fill='x')
                    
                    cancel_btn = ctk.CTkButton(cancel,text='•ENTER',font=('courier',30,'underline','bold'),command=lambda:cancel_function())
                    cancel_btn.pack(fill='x')

                    cancel.mainloop()

#=================================================IMAGE PATH======================================================================
                creditcard_image = ctk.CTkImage(dark_image=Image.open(r"creditcard.png"), size=(180, 120))
                gcash_image = ctk.CTkImage(dark_image=Image.open(r"gcash.png"), size=(180, 120))
                paymaya_image = ctk.CTkImage(dark_image=Image.open(r"paymaya.png"), size=(180, 100))
                armory_image = ctk.CTkImage(dark_image=Image.open(r"Armory.png"), size=(450, 650))
                paybg_image = ctk.CTkImage(dark_image=Image.open(r"paybg.png"), size=(550, 700))

#=================================================FRAME,LABEL,ENTRY,CHECKBOX======================================================================
                go_frame = ctk.CTkFrame(co, width=450, height=650, corner_radius=5)
                go_frame.place(x=0, y=50)

                armorybg_label = ctk.CTkLabel(go_frame, text='',image=armory_image)
                armorybg_label.grid(row=1,column=0,rowspan=10,columnspan=10)

                go_label = ctk.CTkLabel(go_frame, text='Go To Main\nTO CheckOut ', font=('Arial', 50, 'underline', 'bold'))
                go_label.grid(row=0,column=5)

                pay_frame = ctk.CTkFrame(co,width=550,height=700,corner_radius=5)
                pay_frame.place(x=450,y=0)

                paybg_label = ctk.CTkLabel(pay_frame,image=paybg_image,text='')
                paybg_label.place(x=0,y=0)

                credit_var = ctk.StringVar()
                creditchk = ctk.CTkCheckBox(pay_frame,text='Credit Card',variable=credit_var,onvalue='Credit Card',offvalue='',font=('arial',30,'bold'))
                creditchk.place(x=25,y=75)

                creditcardimg = ctk.CTkLabel(pay_frame,text='',image=creditcard_image)
                creditcardimg.place(x=300,y=30)

                gcash_entry = ctk.CTkEntry(pay_frame,placeholder_text='#Gcash Number',font=('arial',30,'bold'),width=300)
                gcash_entry.place(x=25,y=225)

                gcashimg = ctk.CTkLabel(pay_frame,text='',image=gcash_image)
                gcashimg.place(x=360,y=180)

                paymaya_entry = ctk.CTkEntry(pay_frame,placeholder_text='#PayMaya Number',font=('arial',30,'bold'),width=300)
                paymaya_entry.place(x=25,y=375)

                paymayaimg = ctk.CTkLabel(pay_frame,text='',image=paymaya_image,corner_radius=5)
                paymayaimg.place(x=350,y=340)

                username_entry = ctk.CTkEntry(pay_frame,placeholder_text='•UserName',font=('arial',30,'bold'),width=300)
                name_entry = ctk.CTkEntry(pay_frame,placeholder_text='•Full Name',font=('arial',30,'bold'),width=300)
                add_entry = ctk.CTkEntry(pay_frame,placeholder_text='•Address',font=('arial',30,'bold'),width=300)

                username_entry.place(x=25,y=475)
                name_entry.place(x=25,y=525)
                add_entry.place(x=25,y=575)
                
                pay_button = ctk.CTkButton(pay_frame,text='Pay',font=('arial',30,'bold'),fg_color='#323232',command=lambda:show_selected(co_frame))
                pay_button.place(x=370,y=630)

                cancel_button = ctk.CTkButton(pay_frame,text='Cancel Order',font=('arial',30,'bold'),fg_color='#323232',command=lambda:cancel_order())
                cancel_button.place(x=25,y=630)

#=================================================FOR EVERY TIME CHECKOUT INTERFACE CLOSE IT GO BACKS TO THE SHOP INTERFACE======================================================================
                co.protocol("WM_DELETE_WINDOW", lambda: go_back_to_main(shop, co))
                co.mainloop()

#=================================================FOR EVERY TIME SHOP INTERFACE CLOSE IT GO BACKS TO THE LOGIN INTERFACE======================================================================
            shop.protocol("WM_DELETE_WINDOW", lambda: go_back_to_main(root, shop))
            clear_login_entry()
            shop.mainloop()

#=================================================SHOWS IF THE USERNAME OR PASSWORD IS WRONG======================================================================
        else:
            messagebox.showerror("Notification", "Wrong Password / Username!\n Try Again!", parent=root)

#=================================================EXIT FUNCTION======================================================================
def cancelfunction():
    root.destroy()

#=================================================CHANGE PASSWORD FUNCTION======================================================================
def change_password():
    put = ctk.CTkToplevel()
    put.geometry('400x200')
    put.title('CHANGE PASSWORD')
    put.resizable(False,False)
    put.configure(bg='dimgray')
    root.withdraw()

#=================================================ACCOUNTS(EXCEL FOR THE REGISTERED ACCOUNTS)======================================================================
    excel_con = load_workbook('accounts.xlsx')
    excel_activate = excel_con.active

#=================================================TO GO BACK TO LOGIN INTERFACE======================================================================
    def go_back_to_main(root, put):
        put.destroy()
        root.deiconify()
        messagebox.showinfo("Login", "Back To LogIn")
    
#=================================================NEW INTERFACE FOR CHANGING PASSWORD)======================================================================
    def go_to_change():
        change = put_entry.get()
        if change == "" :
            messagebox.showinfo("Notification", "Email required", parent=root)
        else:
            for each_cell in range(2, (excel_activate.max_row)+1):
                if (change == excel_activate['C'+str(each_cell)].value):
                    Found = True
                    break
                else:
                    Found = False
            if (Found == True):
                change = ctk.CTkToplevel()
                change.geometry('600x310')
                change.title('CHANGE PASSWORD')
                change.configure(bg='dimgray')
                put.withdraw()

                def update():
                    passcode = pas_ntr.get()

                    if passcode == "":
                        messagebox.showinfo("Notification", "All fields are required", parent=root)
                    elif len(passcode) < 8:
                        messagebox.showerror("Notification", "Password must be at least 8 characters long!", parent=root)
                    else:
                        excel_activate['B'+str(each_cell)].value = pas_ntr.get()
                        excel_con.save('accounts.xlsx')
                        change.destroy()
                        messagebox.showinfo("Notification","Password Changed Successfully!!!")
                        root.deiconify()
                    
                edit_label = ctk.CTkLabel(change,
                                        text='Input New Password',
                                        font=('arial',30,'underline'))
                edit_label.pack(pady=10)

                pas_label = ctk.CTkLabel(change,
                                        text="New Password :",
                                        pady=20,
                                        font=('arial',20))
                pas_ntr = ctk.CTkEntry(change,
                                    width=200,
                                    show='•',
                                    font=('arial',20),
                                    placeholder_text='•PASSWORD')
                
                pas_label.pack(pady=10)
                pas_ntr.pack(pady=10)

                updatebtn = ctk.CTkButton(change,
                                        width=15,
                                        text='Change Password',
                                        command=lambda:update(),
                                        font=('arial',30),
                                        corner_radius=10,
                                        fg_color='#323232',
                                        hover_color='black')
                updatebtn.pack(pady=10)

                change.mainloop()
            else:
                messagebox.showerror("Notification", "Wrong Email", parent=root)

#=================================================LABEL,ENTRY,BUTTON======================================================================
    put_email = ctk.CTkLabel(put,text='InPut Your Email :',
                            width=20,
                            font=('arial',20,'underline'))
    put_entry = ctk.CTkEntry(put,
                            width=200,
                            font=('arial',20),
                            placeholder_text='•EMAIL')

    put_email.pack(pady=10)
    put_entry.pack(pady=10)

    enter_btn = ctk.CTkButton(put,
                            text="Enter",
                            width=14,command=lambda:go_to_change(),
                            font=('arial',30),
                            fg_color='#323232')
    enter_btn.pack(pady=10)

#=================================================WHENEVER CLOSES IT GOES BACK TO LOGIN INTERFACE======================================================================
    put.protocol("WM_DELETE_WINDOW", lambda: go_back_to_main(root, put))
    put.mainloop()

#=================================================FORGOT PASSWORD FUNCTION======================================================================
def forgot_password():
    forgot = ctk.CTkToplevel()
    forgot.geometry('200x200')
    forgot.resizable(False,False)
    forgot.title('FORGOT PASSWORD')
    forgot.configure(bg='dimgray')
    root.withdraw()

#=================================================ACCOUNTS(EXCEL FOR THE REGISTERED ACCOUNTS)======================================================================
    excel_con = load_workbook('accounts.xlsx')
    excel_activate = excel_con.active

#=================================================TO GO BACK TO LOGIN INTERFACE======================================================================
    def go_back_to_main(root, forgot):
        forgot.destroy()
        root.deiconify()
        messagebox.showinfo("Login", "Back To LogIn")

#=================================================NEW INTERFACE TO CHANGE THE PASSWORD======================================================================
    def go_to_change():
        forgotn = forgot_entry_name.get()
        forgote = forgot_entry.get()
        each_cell = excel_activate.max_row
        if forgotn == "" or forgote == "":
            messagebox.showinfo("Notification", "All fields are required", parent=root)
        elif forgotn == "" and (forgote == excel_activate['C'+str(each_cell)].value):
            messagebox.showinfo("Notification", "All fields are required", parent=root)
        elif (forgotn == excel_activate['A'+str(each_cell)].value) and forgote == "":
            messagebox.showinfo("Notification", "All fields are required", parent=root)
        else:
            for each_cell in range(2, (excel_activate.max_row)+1):
                if (forgot_entry_name.get() == excel_activate['A'+str(each_cell)].value and forgot_entry.get() == excel_activate['C'+str(each_cell)].value):
                    Found = True
                    break
                else:
                    Found = False
            if (Found == True):
                forgot_change = ctk.CTkToplevel()
                forgot_change.geometry('600x210')
                forgot_change.title('FORGOT PASSWORD')
                forgot_change.configure(bg='dimgray')
                forgot_change.resizable(False,False)
                forgot.withdraw()

                def update():
                    passcode = pas_ntr.get()

                    if passcode == "":
                        messagebox.showinfo("Notification", "All fields are required", parent=root)
                    elif len(passcode) < 8:
                        messagebox.showerror("Notification", "Password must be at least 8 characters long!", parent=root)
                    else:
                        excel_activate['B'+str(each_cell)].value = pas_ntr.get()
                        excel_con.save('accounts.xlsx')
                        forgot_change.destroy()
                        messagebox.showinfo("Notification","Password Changed Successfully!!!")
                        root.deiconify()
                    
                edit_label = ctk.CTkLabel(forgot_change,
                                          text='Input New Password',
                                          font=('arial',30,'underline'))
                edit_label.pack(pady=10)

                pas_label = ctk.CTkLabel(forgot_change,
                                        text="New Password :",
                                        font=('arial',20))
                pas_ntr = ctk.CTkEntry(forgot_change,
                                       width=200,
                                       show='•',
                                       font=('arial',20),
                                       placeholder_text='•PASSWORD')
                
                pas_label.pack(pady=10)
                pas_ntr.pack(pady=10)

                updatebtn = ctk.CTkButton(forgot_change,
                                          width=15,
                                          text='Change Password',
                                          command=lambda:update(),
                                          font=('arial',30),
                                          corner_radius=10,
                                          fg_color='#323232',
                                          hover_color='black',
                                          height=2)
                updatebtn.pack(pady=10)

                forgot_change.mainloop()
            
            else:
                messagebox.showerror('NOTIFICATION','WRONG USERNAME/EMAIL \n TRY AGAIN!')

#=================================================LABEL,ENTRY,BUTTON======================================================================
    forgot_name = ctk.CTkLabel(forgot,
                               text='InPut Your Name :', 
                               width=20,
                               font=('arial',20,'underline'))
    forgot_entry_name = ctk.CTkEntry(forgot, 
                                     width=200,
                                     font=('arial',20),
                                     placeholder_text='•USERNAME')

    forgot_name.pack(pady=(10,5))
    forgot_entry_name.pack(pady=5)

    forgot_email = ctk.CTkLabel(forgot,
                                text='InPut Your Email :', 
                                width=20,
                                font=('arial',20,'underline'))
    forgot_entry = ctk.CTkEntry(forgot, 
                                width=200,
                                font=('arial',20),
                                placeholder_text='•EMAIL')

    forgot_email.pack(pady=5)
    forgot_entry.pack(pady=5)

    enter_btn = ctk.CTkButton(forgot,
                              text="Enter",
                              width=14,
                              command=lambda:go_to_change(),
                              font=('arial',30),
                              fg_color="#323232",
                              hover_color='black')
    enter_btn.pack(pady=5)

#=================================================WHENEVER CLOSES IT GOES BACK TO LOGIN INTERFACE======================================================================
    forgot.protocol("WM_DELETE_WINDOW", lambda: go_back_to_main(root, forgot))
    forgot.mainloop()

#=================================================USER \ PASS ENTRY ======================================================================
user_frame =ctk.CTkFrame(root,width=450,height=50,fg_color='#323232')
user_frame.place(x=80,y=250)

pass_frame =ctk.CTkFrame(root,width=450,height=50,fg_color='#323232')
pass_frame.place(x=80,y=300)

userframe_label = ctk.CTkLabel(user_frame, image=user_image, text="")
userframe_label.grid(row=0,column=0) 
passframe_label = ctk.CTkLabel(pass_frame, image=pass_image, text="")
passframe_label.grid(row=0,column=0) 
logoframe_label = ctk.CTkLabel(root, image=logo_image, text="",fg_color='black',corner_radius=10)
logoframe_label.place(x=175,y=50)

username = ctk.CTkEntry(user_frame,
                 width=400,
                 font=('arial',30),
                 placeholder_text='Username',
                 corner_radius=0,)
password = ctk.CTkEntry(pass_frame,
                 width=400,
                 show="•",
                 font=('arial',30),
                 placeholder_text='Password',
                 corner_radius=0)
login = ctk.CTkButton(root,
               text="LOGIN",
               width=450,
               command=lambda:loginfunction(),
               font=('arial',30),
               anchor='CENTER',
               fg_color='#323232',
               text_color='white',
               hover_color='black')
cancel = ctk.CTkButton(root,
                text="EXIT",
                width=20,
                command=lambda:cancelfunction(),
                font=('arial',30),
                anchor='CENTER',
               fg_color='#323232',
               text_color='white',
               hover_color='black')
register = ctk.CTkButton(root,
                  text="Register",
                  command=lambda:registerfunction(),
                  width=16,
                  font=('arial',25,'underline'),
                  fg_color='#323232',
                hover_color='black')
change_pass = ctk.CTkButton(root,
                            text="Change Password",
                            width=14,
                            command=lambda:change_password(),
                            font=('arial',15,'underline'),
                            fg_color='#323232',
                            hover_color='black')
forgot_password_pass = ctk.CTkButton(root,
                            text="Forgot Password",
                            width=14,
                            command=lambda:forgot_password(),
                            font=('arial',15,'underline'),
                            fg_color='#323232',
                            hover_color='black')
brand_name = ctk.CTkLabel(root,
                        text='ArmorerΩ',
                        font=('courier',30,'underline'),
                        text_color='white',fg_color='black')
brand_label = ctk.CTkLabel(root,
                           text='Equip With Our \nLatest Gun Innovation',
                           font=('courier',25,'bold'),fg_color='black',corner_radius=10)
register_label = ctk.CTkLabel(root,
                              text="DON'T HAVE AN ACCOUNT?",
                              font=('arial',27,'bold'))

username.grid(row=0,column=1)
password.grid(row=0,column=1) 
login.place(x=80,
            y=370)
cancel.place(x=260,
             y=550)
register.place(x=460,
               y=438)
change_pass.place(x=150,
                  y=510)
forgot_password_pass.place(x=300,
                  y=510)
brand_name.place(x=280
                 ,y=70)
brand_label.place(x=160,
                  y=150)
register_label.place(x=80,y=440)

root.mainloop()